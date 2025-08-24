# tasks.py
import re
import time
import os
import datetime
from zoneinfo import ZoneInfo
from celery import Celery
from scraper_refactored import UzemScraper 
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill


celery_app = Celery('tasks', broker='redis://redis:6379/0', backend='redis://redis:6379/0')

def create_excel_report(data, minimum_values, task_id):
    """Verilen dataya göre biçimlendirilmiş bir Excel raporu oluşturur ve kaydeder."""
    output_folder = 'output'
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    now = datetime.datetime.now(ZoneInfo("Europe/Istanbul"))
    timestamp = now.strftime("%d-%m-%Y_%H-%M")
    filename = os.path.join(output_folder, f'UZEM_DOYK_{timestamp}.xlsx')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'DOYK Analizi'

    # Stil tanımlamaları
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    warning_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # Sütun genişlikleri
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 5
    sheet.column_dimensions['D'].width = 5
    sheet.column_dimensions['E'].width = 5
    sheet.column_dimensions['F'].width = 5

    current_row = 1
    for language, levels_data in data.items():
        if not levels_data: continue

        # Her dil bloğundan önce boş satır bırak
        if current_row > 1:
            current_row += 1
        
        # Başlıklar
        sheet.cell(row=current_row, column=3, value="DOYK").alignment = center_align
        sheet.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=6)
        
        header_row = current_row + 1
        sheet.cell(row=header_row, column=1, value="Dil").alignment = center_align
        sheet.cell(row=header_row, column=2, value="Seviye").alignment = center_align
        sheet.cell(row=header_row, column=3, value="D").alignment = center_align
        sheet.cell(row=header_row, column=4, value="O").alignment = center_align
        sheet.cell(row=header_row, column=5, value="Y").alignment = center_align
        sheet.cell(row=header_row, column=6, value="K").alignment = center_align

        for col in range(1, 7):
            sheet.cell(row=current_row, column=col).border = thin_border
            sheet.cell(row=header_row, column=col).border = thin_border

        # Veri satırları
        data_start_row = header_row + 1
        min_value = minimum_values.get(language, 42)

        sorted_levels = sorted(levels_data.keys(), key=lambda k: (k[0], int(k[1:])) if len(k) > 1 and k[1:].isdigit() else (k, 0))

        for i, level_code in enumerate(sorted_levels):
            row_num = data_start_row + i
            doyk_counts = levels_data[level_code]
            
            sheet.cell(row=row_num, column=2, value=level_code).alignment = center_align
            
            col_map = {'D': 3, 'O': 4, 'Y': 5, 'K': 6}
            for skill, col_idx in col_map.items():
                cell = sheet.cell(row=row_num, column=col_idx)
                value = doyk_counts.get(skill, 0)
                cell.value = value
                cell.alignment = center_align
                if value < min_value:
                    cell.fill = warning_fill
            
            for col in range(1, 7):
                sheet.cell(row=row_num, column=col).border = thin_border
        
        # Dil adını birleştir ve ortala
        if len(sorted_levels) > 0:
            end_row = data_start_row + len(sorted_levels) - 1
            sheet.merge_cells(start_row=data_start_row, start_column=1, end_row=end_row, end_column=1)
            lang_cell = sheet.cell(row=data_start_row, column=1)
            lang_cell.value = language
            lang_cell.alignment = center_align

        current_row = data_start_row + len(sorted_levels)
    
    workbook.save(filename)
    return filename

@celery_app.task(bind=True)
def start_scrape_process(self, username, password, minimum_values, selected_languages):
    """
    Celery ana görevi:
      1) Giriş + linkleri çek
      2) Hedef seviye sayfalarından kursları topla
      3) TR+EN başlıklardan beceri (D/O/Y/K) tespiti yap
      4) Excel raporunu üret
      5) Tüm uyarı/teşhis mesajlarını HTML log-area'ya da ilet
    """

    # --- Yardımcılar ---
    def update_status(progress, message):
        # Arayüze ilerleme + log mesajı gönder
        self.update_state(state='PROGRESS', meta={'progress': progress, 'log_message': message})
        # Çok hızlı ardışık update'lerde UI'nin yetişebilmesi için küçük gecikme
        time.sleep(0.3)

    def send_log_to_frontend(progress, message):
        # Konsola da yaz, UI'ya da yolla
        print(message)
        update_status(progress, message)

    # TR + EN eşleşme desenleri (başlık içi)
    SKILL_PATTERNS = {
        'D': [r'\bdinleme\b', r'\blistening\b'],
        'O': [r'\bokuma\b', r'\breading\b'],
        'Y': [r'\byazma\b', r'\bwriting\b'],
        'K': [r'\bkonuşma\b', r'\bkonusma\b', r'\bspeaking\b'],
    }
    SKILL_FALLBACK_RE = re.compile(r'\b(reading|writing|listening|speaking)\s+skill(s)?\b', re.IGNORECASE)
    SKILL_WORD_TO_CODE = {'reading': 'O', 'writing': 'Y', 'listening': 'D', 'speaking': 'K'}

    def detect_skill_from_title(title: str):
        t = (title or '').lower()
        for code, pats in SKILL_PATTERNS.items():
            for p in pats:
                if re.search(p, t, re.IGNORECASE):
                    return code
        m = SKILL_FALLBACK_RE.search(t)
        if m:
            return SKILL_WORD_TO_CODE.get(m.group(1).lower())
        return None

    LEVEL_RE = re.compile(r'\b(A1|A2|B1|B2|C1|C2)\b', re.IGNORECASE)

    # Hangi dilde hangi seviyeler işlenecek (gerekirse özelleştir)
    allowed_levels_config = {
        'İngilizce': ['A1', 'A2', 'B1', 'B2', 'C1'],
        'default':   ['A1', 'A2', 'B1']
    }

    scraper = None
    try:
        # 1) Giriş
        update_status(5, 'Sistem başlatılıyor...')
        scraper = UzemScraper(username, password)

        update_status(10, 'WebDriver bağlanıyor...')
        if not scraper.connect_driver():
            raise Exception("WebDriver başlatılamadı.")

        update_status(15, 'Sisteme giriş yapılıyor...')
        if not scraper.login():
            raise Exception("Giriş başarısız. Lütfen kullanıcı adı/şifreyi kontrol edin.")

        update_status(30, 'Dil seviyesi linkleri çekiliyor...')
        language_data = scraper.get_language_level_links()
        if not language_data:
            raise Exception("Ana sayfadan dil seviyesi linkleri çekilemedi.")

        if selected_languages:
            language_data = {lang: levels for lang, levels in language_data.items() if lang in selected_languages}
            if not language_data:
                raise Exception("Seçilen diller için link bulunamadı.")

        # 2) İşlenecek seviye listesi
        levels_to_process = []
        for lang, levels in language_data.items():
            target_levels = allowed_levels_config.get(lang, allowed_levels_config['default'])
            for level_name, level_url in levels.items():
                m = LEVEL_RE.search(level_name or '')
                level_code = (m.group(0).upper() if m else None)
                if level_code and level_code in target_levels:
                    levels_to_process.append({
                        'lang': lang,
                        'level_name': level_name,
                        'level_code': level_code,
                        'level_url': level_url
                    })
                else:
                    send_log_to_frontend(28, f"[SKIP] Hedef dışı seviye: {lang} / '{level_name}' (izinli: {target_levels})")

        total_levels = len(levels_to_process)
        if total_levels == 0:
            raise Exception("İşlenecek uygun seviye bulunamadı.")

        update_status(30, f'Toplam {total_levels} seviye için kurslar toplanacak...')

        # 3) Kursları topla (hafif yöntemlerle)
        all_scraped_courses = {}  # {lang: {level_code: [ {title, url, total_resources_from_js}, ... ]}}
        for idx, item in enumerate(levels_to_process, start=1):
            lang, level_name, level_code, level_url = item['lang'], item['level_name'], item['level_code'], item['level_url']
            prog = 30 + int(50 * (idx-1) / total_levels)
            update_status(prog, f"İşleniyor: {lang} - {level_name} ({level_code})")

            if lang not in all_scraped_courses:
                all_scraped_courses[lang] = {}
            if level_code not in all_scraped_courses[lang]:
                all_scraped_courses[lang][level_code] = []

            courses_in_level = scraper.scrape_doyk_content(level_url) or []
            if not courses_in_level:
                send_log_to_frontend(prog, f"[WARN] Kurs bulunamadı: {lang} - {level_name} ({level_code})")
            else:
                all_scraped_courses[lang][level_code].extend(courses_in_level)

        # 4) Gruplama (D/O/Y/K)
        update_status(82, 'Veriler işleniyor (D/O/Y/K gruplanıyor)...')

        grouped_data_by_language = {}  # {lang: {level_code: {'D':int,'O':int,'Y':int,'K':int}}}
        for lang, levels_dict in all_scraped_courses.items():
            lang_block = {}
            for level_code, courses_list in levels_dict.items():
                doyk_counts = {'D': 0, 'O': 0, 'Y': 0, 'K': 0}
                for c in (courses_list or []):
                    title = c.get('title', '') or ''
                    total = int(c.get('total_resources_from_js', 0) or 0)

                    skill = detect_skill_from_title(title)
                    if skill:
                        # Aynı seviyede aynı beceriye birden fazla kurs varsa topluyoruz
                        doyk_counts[skill] += total
                    else:
                        # Eşleşmeyen başlıkları UI'ya da gönder
                        send_log_to_frontend(85, f"[WARN] Beceri eşleşmedi: {lang} {level_code} → '{title}' (total={total})")

                lang_block[level_code] = doyk_counts

            if lang_block:
                grouped_data_by_language[lang] = lang_block

        # 5) Sürücüyü kapat + Excel
        if scraper:
            scraper.close_driver()
            scraper = None

        update_status(92, 'Excel raporu oluşturuluyor...')
        excel_file_path = create_excel_report(grouped_data_by_language, minimum_values, self.request.id)

        update_status(100, 'Tamamlandı.')
        return {
            'status': 'SUCCESS',
            'data': grouped_data_by_language,
            'excel_filename': os.path.basename(excel_file_path)
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        # Hata ayrıntısını UI'ya da gönder
        self.update_state(state='FAILURE', meta={'progress': 0, 'log_message': f"Hata: {str(e)}"})
        return {'status': 'FAILURE', 'log_message': str(e)}

    finally:
        if scraper and scraper.driver:
            scraper.close_driver()
